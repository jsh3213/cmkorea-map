import subprocess

import numpy as np
import pyautogui as pag
import time
import clipboard
from openpyxl import Workbook
import os
import win32com.client

#
# wb = Workbook()
# file_name = 'sample.xlsx'
#
# ws1 = wb.active
# ws1.title = 'sample excel'
# ws1['A1'] = 'Python excel'
#
# ws2 = wb.create_sheet()
# ws2.title = 'sample excel2'
# ws2['A2'] = 'Python excel'
#
# ws3 = wb.create_sheet()
# ws3.title = 'sample excel3'
# ws3['A3'] = 'Python excel'
#
# wb.save(filename= file_name)





# subprocess.Popen("c:/Wafermap/Wafermap.exe")
# time.sleep(1)
# pag.hotkey('ctrl', 'o')
# pag.typewrite('HARP_3D_149.map', interval=0.05)
# pag.typewrite(['enter'])
# pag.hotkey('alt')
# time.sleep(0.2)
# pag.hotkey('e')
# time.sleep(0.2)
# pag.typewrite(['enter'])
# time.sleep(0.3)
# pag.hotkey('ctrl', 'tab')
# time.sleep(0.2)
# center = pag.locateCenterOnScreen('1.PNG')
# pag.click(center)
# time.sleep(0.2)
# pag.hotkey('ctrl', 'v')
# time.sleep(2)
# center = pag.locateCenterOnScreen('ok.PNG')
# pag.click(center)

#
# edge1 = pag.locateCenterOnScreen('edge1.PNG')
# edge2 = pag.locateCenterOnScreen('edge2.PNG')
# start_edge1, start_edge2 = edge1
# end_edge1, end_edge2 = edge2
# size1 = end_edge1 - start_edge1
# size2 = end_edge2 - start_edge2
# pag.screenshot("sample1.png", region=(start_edge1, start_edge2, size1, size2))

# width, height = pag.size()
# pag.moveTo(width / 2, height / 2)
# pag.moveRel(-154, -142, 0.3)
# time.sleep(0.2)
# pag.click()

import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np
data = np.random.randint(-100, 100, 50).cumsum()
from matplotlib import font_manager, rc

font_name = font_manager.FontProperties(fname="c:/Windows/Fonts/malgun.ttf").get_name()
rc('font', family=font_name)


plt.plot(range(50), data, 'r')
plt.ylabel('가격')
# plt.title('가격변동 추이', fontproperties=fontprop2)
plt.show()

