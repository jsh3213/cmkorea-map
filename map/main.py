import os
import sys, cm
import time
import subprocess
import pyautogui as pag
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QDate
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from PyQt5.QtGui import QIcon
from matplotlib import font_manager, rc



class WindowClass(QMainWindow, cm.Ui_MainWindow):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon('pang.png'))
        self.setWindowTitle("CM KOREA")
        self.date = QDate.currentDate()
        self.statusBar().showMessage(self.date.toString(Qt.DefaultLocaleLongDate))
        self.setupUIxy()

    def setupUIxy(self):
        self.pushButton.clicked.connect(self.getFile)
        self.pushButton_2.clicked.connect(self.setValues)
        self.pushButton_2.setDisabled(True)
        self.make_3d_map.clicked.connect(self.save_3d_map)
        self.make_3d_map.setDisabled(True)
        self.folder_open_button.clicked.connect(self.openFolder)

    def openFolder(self):
        path = os.getcwd() + "/MAP/"
        path = os.path.realpath(path)
        os.startfile(path)

    def createFolder(self):
        dir = os.getcwd()
        try:

            if not os.path.exists(str(dir) + '/MAP'):
                os.makedirs(str(dir) + '/MAP')

            if not os.path.exists(str(dir) + '/3D DATA'):
                os.makedirs(str(dir) + '/3D DATA')
        except OSError:
            print('Error: Creating directory. ' + str(dir) + '/MAP')

    def getFile(self):
        global gValueXY
        global g3dMap

        # gValueXY = [1, 2, 3], [1, 2, 3]
        # g3dMap = [1, 2, 3]
        #
        # gValueXY[0].clear()
        # gValueXY[1].clear()
        # g3dMap.clear()
        # self.tableWidget.clear()

        file = QFileDialog.getOpenFileName()
        fileName = (file[0])

        load_file = fileName[-4:]
        if load_file == 'xlsx':
            heaterType = self.comboBox.currentText()
            if heaterType == 'HARP':
                load_wb = load_workbook(fileName, data_only=True)
                load_ws1 = load_wb['Sheet1']
                load_wb.save(filename='sample.xlsx')

                valueX = []
                valueY = []
                map3d = []
                modifiedValue = 0.01

                for x in range(14, 465, 3):
                    index = "F" + str(x)
                    value = (load_ws1[index].value)
                    valueX.append(round((value * 1000), 2))

                valueX.insert(0, 30)
                valueX.insert(0, 30)
                valueX.insert(0, 30)
                valueX.insert(155, 30)
                valueX.insert(155, 30)
                valueX.insert(155, 30)

                for i in range(0, 157):
                    item = QTableWidgetItem()
                    item.setData(Qt.DisplayRole, valueX[i])
                    self.tableWidget.setItem(i, 0, item)

                # groove line modified - x
                modified = valueX[19]
                valueX[20] = modified - modifiedValue
                modified = valueX[135]
                valueX[136] = modified + modifiedValue

                for x in range(467, 894, 3):
                    index = "F" + str(x)
                    value = (load_ws1[index].value)
                    valueY.append(round((value * 1000), 2))

                valueY.insert(0, 30)
                valueY.insert(0, 30)
                valueY.insert(0, 30)
                valueY.insert(155, 30)
                valueY.insert(155, 30)
                valueY.insert(155, 30)

                modified = valueY[65]
                valueY.insert(66, modified - modifiedValue)
                valueY.insert(67, modified - modifiedValue)
                modified = valueY[86]
                valueY.insert(87, modified + modifiedValue)
                valueY.insert(88, modified + modifiedValue)

                j = valueY[137]
                for i in range(138, 142):
                    j += modifiedValue
                    valueY.insert(i, j)

                # groove line modified - y
                modified = valueY[19]
                valueY[20] = modified - modifiedValue
                modified = valueY[135]
                valueY[136] = modified + modifiedValue
                valueY.reverse()

                for i in range(0, 157):
                    item = QTableWidgetItem()
                    item.setData(Qt.DisplayRole, valueY[i])
                    self.tableWidget.setItem(i, 1, item)

                for item in range(896, 1284, 3):
                    index = "F" + str(item)
                    value = (load_ws1[index].value)
                    map3d.append(round((value * 1000), 2))

                self.pushButton_2.setEnabled(True)

                gValueXY = valueX, valueY
                g3dMap = map3d
            else:
                pass
            # elif heaterType == 'A'

        else:
            QMessageBox.question(self, "확인", 'error', QMessageBox.Ok)

    def modifyValue(self):
        global gValueXY

        item = self.tableWidget.currentItem()
        if not item:
            QMessageBox.question(self, "확인", 'error', QMessageBox.Ok)
        row = self.tableWidget.currentRow()
        column = self.tableWidget.currentColumn()
        self.tableWidget.setItem(row, column, item)
        value = item.text()
        self.changeGvalueXY(row, column, value)

    def changeGvalueXY(self, row, column, value):
        global gValueXY
        gValueXY[column][row] = float(value)

    def setValues(self):
        global gTitleX
        global gTitleY
        global gSavePng
        global gExcelPath

        self.createFolder()
        self.tableWidget.cellChanged.connect(self.modifyValue)


        serialNumber = self.lineEdit.text()
        heaterType = self.comboBox.currentText()
        print(heaterType)
        gTitleX = (serialNumber + '_' + heaterType + '_L-R')
        gTitleY = (serialNumber + '_' + heaterType + '_T-B')
        date = self.date.toString(Qt.ISODate)
        gSavePng = serialNumber + ' ' + heaterType + " " + date
        excelFile = gSavePng + ".xlsx"
        path = 'map/' + gSavePng + '/' + excelFile
        gExcelPath = path


        try:
            if not os.path.exists('map/' + gSavePng):
                os.makedirs('map/' + gSavePng)
        except OSError:
            print('Error: Creating directory. ' + 'map/' + gSavePng)

        wb = load_workbook('sample.xlsx')
        # ws1 = wb.active
        # ws1.title = 'sample excel1'
        ws2 = wb.create_sheet()
        ws2.title = 'DATA'
        ws2['A1'] = 'L -> R'
        ws2['B1'] = 'T -> B'
        ws2['C1'] = '3D DATA'

        numX = 2
        for item in gValueXY[0]:
            changeNum = str(numX)
            numX += 1
            ws2['A' + changeNum] = item

        numY = 2
        for item in gValueXY[1]:
            changeNum = str(numY)
            numY += 1
            ws2['B' + changeNum] = item

        num3D = 2
        for item in g3dMap:
            changeNum = str(num3D)
            num3D += 1
            ws2['C' + changeNum] = item

        wb.save(path)

        left = gValueXY[0][3]
        right = gValueXY[0][-4]
        top = gValueXY[1][3]
        bottom = gValueXY[1][-4]
        min1 = min(gValueXY[0][3:154])
        max1 = max(gValueXY[0][3:154])
        min2 = min(gValueXY[1][3:154])
        max2 = max(gValueXY[1][3:154])



        minMax1 = str(round(max1, 2) - round(min1, 2))
        minMax2 = str(round(max2, 2) - round(min2, 2))

        self.label_max_1.setText("L:" + str(left) + " / " + "R:" + str(right))
        self.label_min_1.setText(str(min1))
        self.label_max_2.setText("T:" + str(top) + " / " + "B:" + str(bottom))
        self.label_min_2.setText(str(min2))
        self.label_result_1.setText(minMax1)
        self.label_result_2.setText(minMax2)

        self.drawGraph()

    def drawGraph(self):

        global gTitleX
        global gTitleY
        global gSavePng
        global gValueXY

        font_name = font_manager.FontProperties(fname="c:/Windows/Fonts/Arial.ttf").get_name()
        rc('font', family=font_name)

        plt.xlim(0, 300)
        plt.ylim(-40, 40)

        plt.figure(figsize=(15, 10))

        ax1 = plt.subplot(2, 1, 1)
        plt.title(gTitleX, position=(0.5, 0.8), fontsize=20)
        ax1.set_xlim(0, 156)
        ax1.set_ylim(-40, 40)
        ax1.spines['bottom'].set_position(('data', 0))
        ax1.spines['bottom'].set_linewidth(0.3)
        ax1.spines['top'].set_visible(False)

        plt.grid(color='#BDBDBD', linestyle='-', linewidth=0.5, alpha=0.9)
        plt.xticks(fontsize=10)
        # plt.xticks(fontsize=10, color='b')

        plt.subplots_adjust(hspace=0, wspace=0)

        ax2 = plt.subplot(2, 1, 2)
        plt.title(gTitleY, position=(0.5, 0.8), fontsize=20)
        plt.axis([155, 0, -40, 40])
        ax2.yaxis.set_ticks_position('right')
        ax2.spines['bottom'].set_position(('data', 0))
        ax2.spines['bottom'].set_linewidth(0.3)
        ax2.spines['top'].set_visible(False)
        plt.grid(color='#BDBDBD', linestyle='-', linewidth=0.5, alpha=0.9)
        plt.xticks(fontsize=10)

        line1 = ax1.plot(gValueXY[0])
        line2 = ax2.plot(gValueXY[1])

        plt.setp(line1, linewidth=3.0)
        plt.setp(line2, linewidth=3.0)
        # plt.setp(line2, color='r', linewidth=3.0)

        font_name = font_manager.FontProperties(fname="c:/Windows/Fonts/malgun.ttf").get_name()
        rc('font', family=font_name)

        fileame = gSavePng + "_XY.jpg"

        plt.savefig('map/' + gSavePng + '/' + fileame, bbox_inches='tight')

        pixmap = QPixmap('map/' + gSavePng + '/' + fileame)
        pixmap = pixmap.scaledToHeight(750)  # 사이즈가 조정
        self.label.setPixmap(pixmap)

        self.make_3d_map.setEnabled(True)

        # plt.show()



    def save_3d_map(self):
        path = os.getcwd() + './' + gExcelPath
        os.startfile(path)

        time.sleep(1.5)
        pag.press('enter')
        time.sleep(0.2)
        pag.hotkey('ctrl', 'pagedown')
        time.sleep(0.2)
        pag.typewrite(['right'])
        time.sleep(0.1)
        pag.typewrite(['right'])
        time.sleep(0.1)
        pag.hotkey('ctrl', 'space')
        # center = pag.locateCenterOnScreen('excel.PNG')
        # pag.click(center)
        time.sleep(0.2)
        pag.hotkey('ctrl', 'c')
        time.sleep(0.2)
        pag.hotkey('alt', 'f4')
        time.sleep(0.2)
        pag.press('enter')
        time.sleep(0.5)

        subprocess.Popen("c:/Wafermap/Wafermap.exe")
        time.sleep(1)
        pag.hotkey('ctrl', 'o')
        time.sleep(0.2)
        pag.typewrite('HARP_3D_130P_200728.map', interval=0.05)
        time.sleep(0.2)
        pag.press('enter')
        time.sleep(0.2)
        pag.press('enter')
        pag.hotkey('alt')
        time.sleep(0.2)
        pag.hotkey('e')
        time.sleep(0.2)
        pag.press('enter')
        time.sleep(0.3)
        pag.press('right')
        time.sleep(0.3)

        # pag.press('tab')
        # time.sleep(0.2)
        # pag.press('right', presses=2, interval=0.2)
        # time.sleep(0.2)
        # pag.keyDown('shift')
        # time.sleep(0.1)
        # pag.press('down', presses=130, interval=0.1)
        # time.sleep(0.2)
        # pag.keyUp('shift')

        width, height = pag.size()
        pag.moveTo(width / 2, height / 2)
        time.sleep(0.2)
        pag.moveRel(-130, -155, 0.3)
        time.sleep(0.2)
        pag.click()
        time.sleep(0.2)

        pag.hotkey('ctrl', 'v')
        time.sleep(0.5)
        pag.moveRel(270, 350, 0.3)
        time.sleep(0.2)
        pag.click()
        time.sleep(0.3)
        pag.moveTo(width / 2, height / 2)
        pag.moveRel(25, -180, 0.3)
        time.sleep(0.2)
        pag.click()
        time.sleep(0.2)

        pag.moveTo(width / 2, height / 2)
        pag.moveRel(-154, -142, 0.3)
        time.sleep(0.2)
        x, y = pag.position()
        time.sleep(0.2)

        path = './map/' + gSavePng + '/'
        end_name = gSavePng + '_3D.jpg'
        saveFile = path + end_name
        pag.screenshot(saveFile, region=(x, y, 440, 300))
        time.sleep(0.2)
        pag.hotkey('alt', 'f4')

        time.sleep(0.2)


        pixmap = QPixmap(saveFile)
        pixmap = pixmap.scaledToHeight(750)  # 사이즈가 조정
        self.label.setPixmap(pixmap)

        pag.alert(text='3D MAP 완료', title='완료', button='OK')



    def closeEvent(self, QCloseEvent):
        ans = QMessageBox.question(self, "종료 확인", '종료하시겠습니까?',
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans == QMessageBox.Yes:
            QCloseEvent.accept()
        else:
            QCloseEvent.ignore()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
